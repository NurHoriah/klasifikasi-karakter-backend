import os
import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from pymongo import MongoClient
from pymongo.errors import DuplicateKeyError
import joblib
import numpy as np
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
from io import StringIO, BytesIO
import csv
from datetime import datetime, timedelta, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from google.auth.transport import requests
from google.oauth2 import id_token
import re
import dns.resolver  # Tambahkan library untuk cek MX record (perlu: pip install dnspython)
import random  # for OTP generation
import string  # for OTP generation
from flask_mail import Mail, Message  # for sending emails
import smtplib  # Tambahkan import smtplib untuk menangani error spesifik autentikasi

# Load environment variables
load_dotenv()  # Simplified env loading to default to the current directory for better reliability in folder-based projects

client = MongoClient(os.getenv("MONGO_URI"))

app = Flask(__name__)

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')  # Email Anda
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')  # App Password Anda
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('MAIL_USERNAME')

# Check if mail configuration is set
if not app.config['MAIL_USERNAME'] or not app.config['MAIL_PASSWORD']:
    print("[ERROR] Mail configuration is not set!")

mail = Mail(app)

# Using only flask-cors CORS() configuration which properly handles all requests

CORS(app, resources={
    r"/api/*": {
        "origins": ["http://localhost:3000", "http://127.0.0.1:3000", "http://localhost:8000", "http://127.0.0.1:8000"],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "supports_credentials": True,
        "max_age": 3600
    },
    r"/predict": {
        "origins": ["http://localhost:3000", "http://127.0.0.1:3000", "http://localhost:8000", "http://127.0.0.1:8000"],
        "methods": ["POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "supports_credentials": True,
        "max_age": 3600
    }
})

# JWT Configuration
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
app.config['JWT_HEADER_TYPE'] = ''
jwt = JWTManager(app)

# MongoDB Configuration
MONGO_URI = os.getenv('MONGO_URI')
print(f"[DEBUG] MONGO_URI: {MONGO_URI}")
if not MONGO_URI:
    print("[ERROR] MONGO_URI tidak di-set!")

try:
    client = MongoClient(MONGO_URI)
    print("[SUCCESS] Connected to MongoDB Atlas")
except Exception as e:
    print(f"[ERROR] MongoDB connection failed: {e}")
    
db = client['system_klasifikasi']
users_collection = db['users']
results_collection = db['hasil']

try:
    users_collection.create_index('email', unique=True)
except Exception as e:
    print(f"[WARNING] Index creation error: {e}")

# Load SVM Model
model = None
try:
    model = joblib.load('../scripts/model.joblib')
    print(f"[SUCCESS] Model loaded successfully. Classes: {model.classes_}")
except Exception as e:
    print(f"[WARNING] Model tidak ditemukan: {e}")

# ===========================
# AUTH ENDPOINTS
# ===========================

def is_valid_email(email):
    """Strict email validation - reject all fake emails and check domain existence"""
    # Check minimum length and format
    if len(email) < 5 or email.count('@') != 1:
        return False
    
    local_part, domain_part = email.split('@')
    
    blacklisted_domains = [
        'mailinator.com', 'tempmail.com', 'guerrillamail.com', 
        '10minutemail.com', 'test.com', 'example.com'
    ]
    if domain_part.lower() in blacklisted_domains:
        return False

    # Validate local part (before @)
    if not local_part or not domain_part:
        return False
    if len(local_part) < 2:
        return False
    
    # No consecutive dots, no leading/trailing dots or special chars
    if '..' in local_part or local_part.startswith('.') or local_part.endswith('.'):
        return False
    if local_part.startswith('-') or local_part.endswith('-'):
        return False
    
    # Validate domain part (after @)
    if '.' not in domain_part:  # Must have at least one dot
        return False
    if domain_part.startswith('.') or domain_part.endswith('.'):
        return False
    if domain_part.startswith('-') or domain_part.endswith('-'):
        return False
    if '..' in domain_part or '.-' in domain_part or '-.' in domain_part:
        return False
    
    # Check domain extension (TLD)
    parts = domain_part.split('.')
    if len(parts) < 2:
        return False
    
    tld = parts[-1]
    if len(tld) < 2:  # TLD must be at least 2 chars
        return False
    if not tld.isalpha():  # TLD must be only letters
        return False
    
    # Allowed characters in local and domain parts
    allowed_local = set('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._%-+')
    allowed_domain = set('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789.-')
    
    if not all(c in allowed_local for c in local_part):
        return False
    if not all(c in allowed_domain for c in domain_part):
        return False
    
    try:
        # Menambahkan timeout 5 detik agar tidak macet (Error 500) jika koneksi lambat
        resolver = dns.resolver.Resolver()
        resolver.lifetime = 5
        resolver.timeout = 5
        resolver.resolve(domain_part, 'MX')
        return True
    except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer, Exception) as e:
        print(f"[v0] DNS Validation failed for {domain_part}: {e}")
        return False

def generate_otp():
    """Generate a 6-digit numeric OTP"""
    return ''.join(random.choices(string.digits, k=6))

@app.route('/api/register', methods=['POST', 'OPTIONS'])
def register():
    """Register user dengan email dan password + kirim OTP"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '').strip()

        if not name or not email or not password:
            return jsonify({'message': 'Nama, email, dan password wajib diisi!'}), 400

        print(f"[v0] Validating email: {email}")
        if not is_valid_email(email):
            print(f"[v0] Email validation failed for: {email}")
            return jsonify({'message': 'Email tidak valid atau domain tidak terdaftar! Gunakan email asli (Gmail/Yahoo/dsb).'}), 400

        existing_user = users_collection.find_one({'email': email})
        if existing_user:
            if not existing_user.get('is_verified', False):
                otp = generate_otp()
                users_collection.update_one(
                    {'email': email},
                    {'$set': {'otp': otp, 'otp_created_at': datetime.now(timezone.utc)}}
                )
                try:
                    print(f"[v0] Attempting to send OTP to {email}")
                    msg = Message('Verifikasi Akun Anda', recipients=[email])
                    msg.body = f'Halo {name},\n\nKode OTP Anda adalah: {otp}\n\nKode ini berlaku selama 5 menit.'
                    mail.send(msg)
                    print("[v0] OTP sent successfully")
                    return jsonify({'message': 'User sudah terdaftar namun belum diverifikasi. OTP baru telah dikirim!', 'needs_verification': True, 'email': email}), 200
                except smtplib.SMTPAuthenticationError:
                    print(f"[v0] ERROR: Password SMTP (App Password) salah!")
                    return jsonify({'message': 'Gagal login ke server email. Periksa apakah MAIL_PASSWORD di .env sudah benar (tanpa spasi).'}), 500
                except Exception as e:
                    print(f"[v0] Mail send error: {str(e)}")
                    return jsonify({'message': f'Gagal mengirim email verifikasi: {str(e)}'}), 500
            
            return jsonify({'message': 'Email sudah terdaftar!'}), 400

        hashed_password = generate_password_hash(password)
        otp = generate_otp()

        user_data = {
            'name': name,
            'email': email,
            'password': hashed_password,
            'role': 'guru',
            'is_verified': False,
            'otp': otp,
            'otp_created_at': datetime.now(timezone.utc),
            'created_at': datetime.now(timezone.utc)
        }
        user_id = users_collection.insert_one(user_data).inserted_id

        try:
            print(f"[v0] Attempting to send OTP to {email}")
            msg = Message('Verifikasi Akun Anda', recipients=[email])
            msg.body = f'Halo {name},\n\nKode OTP Anda adalah: {otp}\n\nKode ini berlaku selama 5 menit.'
            mail.send(msg)
            print("[v0] OTP sent successfully")
        except smtplib.SMTPAuthenticationError:
            print(f"[v0] ERROR: Password SMTP (App Password) salah!")
            users_collection.delete_one({"_id": user_id})
            return jsonify({'message': 'Konfigurasi MAIL_PASSWORD di .env ditolak oleh Google. Pastikan kodenya benar dan tanpa spasi.'}), 500
        except Exception as e:
            print(f"[v0] ERROR CRITICAL: Gagal kirim email: {str(e)}")
            users_collection.delete_one({"_id": user_id})
            return jsonify({
                'message': f'Gagal mengirim email verifikasi. Error: {str(e)}'
            }), 500

        return jsonify({
            'message': 'Pendaftaran berhasil! Silakan cek email Anda untuk kode OTP.',
            'needs_verification': True,
            'email': email
        }), 201

    except Exception as e:
        print(f"[v0] Register Error: {e}")
        return jsonify({'message': 'Terjadi kesalahan pada server'}), 500

@app.route('/api/verify-otp', methods=['POST'])
def verify_otp():
    """Verify email OTP"""
    try:
        data = request.get_json()
        email = data.get('email')
        otp = data.get('otp')

        user = users_collection.find_one({'email': email, 'otp': otp})
        
        if not user:
            return jsonify({'message': 'Kode OTP salah!'}), 400
        
        # Cek kadaluarsa OTP (5 menit)
        otp_time = user.get('otp_created_at')
        if datetime.now(timezone.utc) - otp_time.replace(tzinfo=timezone.utc) > timedelta(minutes=5):
            return jsonify({'message': 'Kode OTP sudah kadaluarsa!'}), 400

        users_collection.update_one(
            {'email': email},
            {'$set': {'is_verified': True}, '$unset': {'otp': "", 'otp_created_at': ""}}
        )

        return jsonify({'message': 'Email berhasil diverifikasi! Sekarang Anda bisa login.'}), 200
    except Exception as e:
        return jsonify({'message': f'Error: {str(e)}'}), 500

@app.route('/api/login', methods=['POST', 'OPTIONS'])
def login():
    """Login user dan return JWT token"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        password = data.get('password', '').strip()

        if not email or not password:
            return jsonify({'message': 'Email dan password wajib diisi!'}), 400

        if not is_valid_email(email):
            return jsonify({'message': 'Format email tidak valid atau domain tidak ada!'}), 400

        user = users_collection.find_one({'email': email})
        if not user:
            return jsonify({'message': 'Email tidak terdaftar! Silakan daftar terlebih dahulu.'}), 404

        if not check_password_hash(user['password'], password):
            return jsonify({'message': 'Password yang Anda masukkan salah!'}), 401

        if not user.get('is_verified', False):
            return jsonify({'message': 'Akun Anda belum diverifikasi! Silakan cek email.', 'needs_verification': True}), 403

        access_token = create_access_token(identity=email)

        return jsonify({
            'message': 'Login berhasil!',
            'token': access_token,
            'user': {
                'id': str(user['_id']),
                'name': user['name'],
                'email': user['email'],
                'picture': user.get('picture', ''),
                'role': user.get('role', 'guru')
            }
        }), 200

    except Exception as e:
        print(f"[ERROR] Login: {e}")
        return jsonify({'message': 'Terjadi kesalahan pada server'}), 500


@app.route('/api/google-login', methods=['POST', 'OPTIONS'])
def google_login():
    """Login user menggunakan Google Sign-In token - ONLY untuk user yang sudah terdaftar"""
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add("Access-Control-Allow-Origin", request.headers.get('Origin', 'http://localhost:3000'))
        response.headers.add("Access-Control-Allow-Methods", "POST, OPTIONS")
        response.headers.add("Access-Control-Allow-Headers", "Content-Type, Authorization")
        response.headers.add("Access-Control-Allow-Credentials", "true")
        response.headers.add("Access-Control-Max-Age", "3600")
        return response, 200
    
    try:
        data = request.get_json()
        token = data.get('token')

        if not token:
            return jsonify({'message': 'Token tidak ditemukan!'}), 400

        try:
            idinfo = id_token.verify_oauth2_token(
                token, 
                requests.Request(), 
                '842122693668-ct0che0nuegv2lk6a5tetv765ib3969o.apps.googleusercontent.com'
            )

            # Verify that the token was issued to us
            if idinfo['aud'] != '842122693668-ct0che0nuegv2lk6a5tetv765ib3969o.apps.googleusercontent.com':
                raise ValueError('Token audience mismatch')

            print(f"[DEBUG] Token verified for: {idinfo.get('email')}")

            email = idinfo.get('email')
            name = idinfo.get('name', 'Google User')
            picture = idinfo.get('picture', '')

            user = users_collection.find_one({'email': email})
            
            if not user:
                # User tidak terdaftar - minta regis dulu
                print(f"[INFO] Google login attempt for unregistered email: {email}")
                return jsonify({'message': 'Email tidak terdaftar! Silakan daftar terlebih dahulu.'}), 404
            
            # User sudah terdaftar, lakukan login
            print(f"[SUCCESS] Existing user logged in with Google: {email}")
            
            # Generate JWT token
            access_token = create_access_token(identity=str(user['_id']))

            return jsonify({
                'message': 'Login dengan Google berhasil!',
                'token': access_token,
                'user': {
                    'id': str(user['_id']),
                    'name': user['name'],
                    'email': user['email'],
                    'picture': user.get('picture', ''),
                    'role': user.get('role', 'guru')
                }
            }), 200

        except ValueError as e:
            print(f"[ERROR] Token verification failed: {e}")
            return jsonify({'message': f'Token tidak valid atau expired!'}), 401

    except ImportError as e:
        print(f"[ERROR] Missing library: {e}")
        print("[ERROR] Install dengan: pip install google-auth")
        return jsonify({'message': 'Server error: google-auth library required'}), 500
    except Exception as e:
        print(f"[ERROR] Google login error: {e}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'message': f'Terjadi kesalahan: {str(e)}'}), 500


# ===========================
# PREDICTION ENDPOINT
# ===========================

@app.route('/predict', methods=['OPTIONS'])
def predict_options():
    """Handle CORS preflight request untuk /predict"""
    response = jsonify({'status': 'ok'})
    response.headers.add("Access-Control-Allow-Origin", "http://localhost:3000")
    response.headers.add("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
    response.headers.add("Access-Control-Allow-Headers", "Content-Type, Authorization")
    response.headers.add("Access-Control-Allow-Credentials", "true")
    response.headers.add("Access-Control-Max-Age", "3600")
    return response, 200

@app.route('/predict', methods=['POST'])
@jwt_required()
def predict():
    """Prediksi tipe karakter belajar"""
    try:
        if model is None:
            return jsonify({'error': 'Model belum di-load'}), 500

        current_user = get_jwt_identity()
        data = request.get_json()

        nama_siswa = data.get('nama_siswa', '').strip()
        kelas = data.get('kelas', '').strip()
        if not nama_siswa or not kelas:
            return jsonify({'error': 'nama_siswa dan kelas harus diisi!'}), 400

        try:
            fitur = [
                float(data.get('nilai_bahasa') or 0),
                float(data.get('nilai_mtk') or 0),
                float(data.get('nilai_ipa') or 0),
                float(data.get('nilai_ips') or 0),
                float(data.get('rata_rata_umum') or 0),
                float(data.get('indeks_eksakta') or 0),
                float(data.get('indeks_non_eksakta') or 0),
                float(data.get('daya_visual_gambar') or 0),
                float(data.get('mengingat_suara') or 0),
                float(data.get('suka_praktik') or 0),
                float(data.get('suka_membaca_mencatat') or 0),
                float(data.get('ekskul_motorik') or 0),
                float(data.get('ekskul_musik') or 0),
            ]
        except (ValueError, TypeError) as e:
            print(f"[ERROR] Feature extraction error: {e}")
            return jsonify({'error': f'Invalid feature values: {str(e)}'}), 400

        if len(fitur) != 13:
            return jsonify({'error': f'Expected 13 features, got {len(fitur)}'}), 400

        fitur_array = np.array([fitur], dtype=float)

        raw_prediction = model.predict(fitur_array)[0]
        probabilities = model.predict_proba(fitur_array)[0]

        prediction_type = str(raw_prediction)
        
        probs_dict = {}
        if hasattr(model, 'classes_'):
            for i, class_name in enumerate(model.classes_):
                if i < len(probabilities):
                    probs_dict[str(class_name)] = float(probabilities[i])
        else:
            probs_dict = {prediction_type: float(max(probabilities))}

        result_data = {
            'user_id': current_user,
            'nama_siswa': nama_siswa,
            'kelas': kelas,
            'student_data': data,
            'prediction': prediction_type,
            'confidence': float(max(probabilities)),
            'probabilities': probs_dict,
            'timestamp': datetime.now(timezone.utc)
        }
        
        insert_result = results_collection.insert_one(result_data)

        tips_map = {
            'Visual': [
                'Gunakan mind map dan highlight warna untuk memudahkan pemahaman',
                'Tampilkan materi dalam bentuk diagram, peta konsep, atau infografis',
                'Rangkum poin penting dengan simbol, warna, atau gambar yang menarik',
                'Ciptakan catatan terstruktur dengan layout yang rapi dan visual'
            ],
            'Auditori': [
                'Dengarkan penjelasan guru, audiobook, atau video pembelajaran dengan seksama',
                'Gunakan lagu, ritme, atau jingle untuk mengingat konsep-konsep penting',
                'Diskusikan pelajaran dengan teman atau kelompok belajar',
                'Merekam penjelasan guru dan dengarkan kembali saat belajar'
            ],
            'Kinestetik': [
                'Lakukan eksperimen langsung, praktik, atau demonstrasi untuk setiap materi',
                'Gunakan alat peraga edukatif, model, atau benda nyata saat belajar',
                'Selingi belajar dengan aktivitas fisik, gerakan, atau permainan edukatif',
                'Buatlah proyek atau karya tangan untuk menerapkan konsep yang dipelajari'
            ]
        }
        
        explanation_map = {
            'Visual': f'Siswa {nama_siswa} adalah tipe belajar Visual. Tipe ini belajar paling efektif melalui gambar, warna, diagram, dan informasi visual lainnya.',
            'Auditori': f'Siswa {nama_siswa} adalah tipe belajar Auditori. Tipe ini belajar paling efektif melalui suara, diskusi, mendengarkan, dan penjelasan lisan.',
            'Kinestetik': f'Siswa {nama_siswa} adalah tipe belajar Kinestetik. Tipe ini belajar paling efektif melalui praktik langsung, eksperimen, dan aktivitas bergerak.'
        }

        response_data = {
            'message': 'Hasil klasifikasi tersimpan',
            'prediction': prediction_type,
            'label': prediction_type,
            'confidence': float(max(probabilities)),
            'probabilities': probs_dict,
            'result_id': str(insert_result.inserted_id),
            'explanation': explanation_map.get(prediction_type, f'Siswa {nama_siswa} adalah tipe belajar {prediction_type}.'),
            'tips': tips_map.get(prediction_type, [
                'Tips umum untuk tipe belajar ini',
                'Silakan konsultasikan dengan guru untuk strategi belajar yang lebih spesifik'
            ])
        }
        
        return jsonify(response_data), 200

    except Exception as e:
        print(f"[ERROR] Predict endpoint: {e}")
        return jsonify({'error': str(e)}), 500


# ===========================
# DATA MANAGEMENT ENDPOINTS
# ===========================

@app.route('/api/results', methods=['GET'])
@jwt_required()
def get_results():
    """Ambil semua hasil prediksi user"""
    try:
        current_user = get_jwt_identity()
        results = list(results_collection.find({'user_id': current_user}))
        
        for result in results:
            result['_id'] = str(result['_id'])
            result['user_id'] = str(result['user_id'])
        
        return jsonify({'results': results}), 200
    except Exception as e:
        print(f"[ERROR] Get results: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/results-filtered', methods=['POST'])
@jwt_required()
def get_results_filtered():
    """Ambil hasil prediksi dengan filter berdasarkan tanggal dan kelas"""
    try:
        current_user = get_jwt_identity()
        data = request.get_json()
        
        filter_query = {'user_id': current_user}
        
        # Filter by date (hari spesifik)
        if data.get('date'):
            date_str = data.get('date')
            try:
                filter_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                start_datetime = datetime.combine(filter_date, datetime.min.time())
                end_datetime = datetime.combine(filter_date, datetime.max.time())
                
                filter_query['timestamp'] = {
                    '$gte': start_datetime,
                    '$lte': end_datetime
                }
            except Exception as e:
                print(f"[ERROR] Date parsing error: {e}")
                pass
        
        # Filter by date range
        if data.get('date_from') and data.get('date_to'):
            try:
                date_from = datetime.strptime(data.get('date_from'), '%Y-%m-%d').date()
                date_to = datetime.strptime(data.get('date_to'), '%Y-%m-%d').date()
                
                start_datetime = datetime.combine(date_from, datetime.min.time())
                end_datetime = datetime.combine(date_to, datetime.max.time())
                
                filter_query['timestamp'] = {
                    '$gte': start_datetime,
                    '$lte': end_datetime
                }
            except Exception as e:
                print(f"[ERROR] Date range parsing error: {e}")
                pass
        
        # Filter by class
        if data.get('kelas'):
            filter_query['kelas'] = data.get('kelas')
        
        results = list(results_collection.find(filter_query))
        
        for result in results:
            result['_id'] = str(result['_id'])
            result['user_id'] = str(result['user_id'])
        
        return jsonify({'results': results, 'count': len(results)}), 200
    except Exception as e:
        print(f"[ERROR] Get filtered results: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/download-all', methods=['GET', 'POST', 'OPTIONS'])
@jwt_required()
def download_all():
    """Download semua hasil prediksi sebagai CSV atau Excel dengan optional filter"""
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add("Access-Control-Allow-Origin", request.headers.get('Origin', 'http://localhost:3000'))
        response.headers.add("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        response.headers.add("Access-Control-Allow-Headers", "Content-Type, Authorization")
        response.headers.add("Access-Control-Allow-Credentials", "true")
        return response, 200

    try:
        current_user = get_jwt_identity()
        
        # Build filter query
        filter_query = {'user_id': current_user}
        
        if request.method == 'POST':
            data = request.get_json() or {}
            
            if data.get('date'):
                try:
                    filter_date = datetime.strptime(data.get('date'), '%Y-%m-%d').date()
                    start_datetime = datetime.combine(filter_date, datetime.min.time())
                    end_datetime = datetime.combine(filter_date, datetime.max.time())
                    filter_query['timestamp'] = {'$gte': start_datetime, '$lte': end_datetime}
                except Exception as e:
                    print(f"[ERROR] Date parsing: {e}")
            
            if data.get('date_from') and data.get('date_to'):
                try:
                    date_from = datetime.strptime(data.get('date_from'), '%Y-%m-%d').date()
                    date_to = datetime.strptime(data.get('date_to'), '%Y-%m-%d').date()
                    start_datetime = datetime.combine(date_from, datetime.min.time())
                    end_datetime = datetime.combine(date_to, datetime.max.time())
                    filter_query['timestamp'] = {'$gte': start_datetime, '$lte': end_datetime}
                except Exception as e:
                    print(f"[ERROR] Date range parsing: {e}")
            
            if data.get('kelas'):
                filter_query['kelas'] = data.get('kelas')
            
            format_type = data.get('format', 'csv').lower()
        else:
            format_type = request.args.get('format', 'csv').lower()
        
        results = list(results_collection.find(filter_query))
        
        if format_type == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Hasil Klasifikasi"
            
            headers = ['Nama Siswa', 'Kelas', 'Prediksi', 'Confidence', 'Tanggal']
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row_num, result in enumerate(results, 2):
                ws.cell(row=row_num, column=1, value=result.get('nama_siswa', ''))
                ws.cell(row=row_num, column=2, value=result.get('kelas', ''))
                ws.cell(row=row_num, column=3, value=result.get('prediction', ''))
                ws.cell(row=row_num, column=4, value=round(result.get('confidence', 0) * 100, 2))
                ws.cell(row=row_num, column=5, value=result.get('timestamp', '').strftime('%Y-%m-%d %H:%M:%S') if result.get('timestamp') else '')
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='hasil_klasifikasi.xlsx')
        
        else:  # CSV format
            output = StringIO()
            writer = csv.writer(output)
            
            writer.writerow(['Nama Siswa', 'Kelas', 'Prediksi', 'Confidence', 'Tanggal'])
            for result in results:
                writer.writerow([
                    result.get('nama_siswa', ''),
                    result.get('kelas', ''),
                    result.get('prediction', ''),
                    round(result.get('confidence', 0) * 100, 2),
                    result.get('timestamp', '').strftime('%Y-%m-%d %H:%M:%S') if result.get('timestamp') else ''
                ])
            
            return send_file(BytesIO(output.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True, download_name='hasil_klasifikasi.csv')
    
    except Exception as e:
        print(f"[ERROR] Download all: {e}")
        return jsonify({'error': str(e)}), 500


# ===========================
# SERVER RUN
# ===========================

if __name__ == '__main__':
    app.run(debug=True, port=8000, host='0.0.0.0')
